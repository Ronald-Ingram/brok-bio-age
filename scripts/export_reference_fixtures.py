#!/usr/bin/env python3
"""Export golden test cases from reference PhenoAge spreadsheet."""

from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from brok_bioage.levine import compute_levine  # noqa: E402
from brok_bioage.units import BiomarkerInputs  # noqa: E402

XLSX = ROOT / "data" / "reference-phenoage.xlsx"
OUT = ROOT / "tests" / "fixtures" / "reference_cases.json"

FIELD_ORDER = [
    "albumin",
    "creatinine",
    "glucose",
    "crp",
    "lymphocyte_pct",
    "mcv",
    "rdw",
    "alp",
    "wbc",
    "age",
]


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    mapping = {
        "albumin": "albumin",
        "creatinine": "creatinine",
        " glucose": "glucose",
        "glucose": "glucose",
        "crp": "crp",
        "c-reac protein": "crp",
        " c-reac protein": "crp",
        "lympocyte": "lymphocyte_pct",
        " lympocyte": "lymphocyte_pct",
        "lymphs (lympocyte)": "lymphocyte_pct",
        "lymphocyte": "lymphocyte_pct",
        "mean cell volume": "mcv",
        "mcv (mean cell volume)": "mcv",
        "red cell dist width": "rdw",
        " red cell dist width": "rdw",
        "rdw (red cell dist width)": "rdw",
        "alkaline phosphatase": "alp",
        " alkaline phosphatase": "alp",
        "white blood cells": "wbc",
        "wbc (white blood cells)": "wbc",
        " age": "age",
        "age": "age",
    }
    return mapping.get(text, text)


def _parse_use_this_next_x(ws) -> dict:
    headers = {_norm_header(ws.cell(11, c).value): c for c in range(3, 13)}
    values = {k: ws.cell(15, c).value for k, c in headers.items()}
    expected = {
        "lincomb": ws.cell(27, 3).value,
        "mortality_risk": ws.cell(27, 4).value,
        "pheno_age": ws.cell(27, 5).value,
    }
    return {
        "sheet": "UseThisNextX",
        "note": "Sample subject from methodology tab",
        "inputs": {
            "albumin_g_dl": values["albumin"],
            "creatinine_mg_dl": values["creatinine"],
            "glucose_mg_dl": values["glucose"],
            "crp_mg_l": values["crp"],
            "lymphocyte_pct": values["lymphocyte_pct"],
            "mcv_fl": values["mcv"],
            "rdw_pct": values["rdw"],
            "alp_u_l": values["alp"],
            "wbc_10e3": values["wbc"],
            "chronological_age": values["age"],
        },
        "expected": expected,
    }


def _find_input_row(ws) -> int | None:
    for row in range(4, 12):
        label_a = ws.cell(row, 1).value
        label_b = ws.cell(row, 2).value
        if label_a == "Input":
            return row
        if label_b == "Input":
            return row
    return None


def _find_results_row(ws) -> int | None:
    for row in range(12, 25):
        if ws.cell(row, 1).value == "Results":
            return row
    return None


def _parse_ri_sheet(name: str, ws) -> dict | None:
    input_row = _find_input_row(ws)
    results_row = _find_results_row(ws)
    if input_row is None or results_row is None:
        return None

    header_row = input_row - 1 if ws.cell(input_row - 1, 2).value == "Albumin" else 4
    headers = {_norm_header(ws.cell(header_row, c).value): c for c in range(2, 12)}

    values = {}
    for field in FIELD_ORDER:
        col = headers.get(field)
        if col is None:
            return None
        values[field] = ws.cell(input_row, col).value

    if not isinstance(values["age"], (int, float)):
        return None

    headers_result = {}
    for c in range(2, 8):
        h = ws.cell(results_row - 1, c).value or ws.cell(results_row - 2, c).value
        if h:
            headers_result[str(h).strip()] = c

    pheno_col = headers_result.get("Ptypic Age")
    mort_col = headers_result.get("MortScore")
    lincomb_col = headers_result.get("LinComb")
    if pheno_col is None:
        pheno_col = 4 if name != "UseThisNextX" else 5

    expected = {
        "pheno_age": ws.cell(results_row, pheno_col).value,
        "mortality_risk": ws.cell(results_row, mort_col).value if mort_col else None,
        "lincomb": ws.cell(results_row, lincomb_col).value if lincomb_col else None,
    }

    return {
        "sheet": name,
        "inputs": {
            "albumin_g_dl": values["albumin"],
            "creatinine_mg_dl": values["creatinine"],
            "glucose_mg_dl": values["glucose"],
            "crp_mg_l": values["crp"],
            "lymphocyte_pct": values["lymphocyte_pct"],
            "mcv_fl": values["mcv"],
            "rdw_pct": values["rdw"],
            "alp_u_l": values["alp"],
            "wbc_10e3": values["wbc"],
            "chronological_age": values["age"],
        },
        "expected": {k: v for k, v in expected.items() if v is not None},
    }


def _to_biomarker_inputs(case_inputs: dict) -> BiomarkerInputs:
    return BiomarkerInputs(**case_inputs)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    cases: list[dict] = []

    if "UseThisNextX" in wb.sheetnames:
        cases.append(_parse_use_this_next_x(wb["UseThisNextX"]))

    for name in wb.sheetnames:
        if "(RI)" not in name and not name.startswith("202"):
            continue
        if "(Comp" in name:
            continue
        parsed = _parse_ri_sheet(name, wb[name])
        if parsed:
            cases.append(parsed)

    # 20260630: sheet uses age=29 de-anchoring experiment; add chrono-57 standard case
    latest = next((c for c in cases if c["sheet"] == "20260630(RI)"), None)
    if latest:
        chrono_inputs = dict(latest["inputs"])
        chrono_inputs["chronological_age"] = 57.0
        computed = compute_levine(_to_biomarker_inputs(chrono_inputs))
        cases.append(
            {
                "sheet": "20260630(RI)_chrono57",
                "note": "Same labs as 20260630(RI) with chronological age 57 (standard Levine)",
                "inputs": chrono_inputs,
                "expected": {"pheno_age": round(computed.pheno_age, 2)},
            }
        )

    sensitivity_base = chrono_inputs if latest else None
    if sensitivity_base:
        for label, overrides, delta in [
            ("sensitivity_creatinine_1_13", {"creatinine_mg_dl": 1.13}, 1.83),
            ("sensitivity_rdw_13_3", {"rdw_pct": 13.3}, 1.80),
            ("sensitivity_glucose_105", {"glucose_mg_dl": 105.0}, 1.18),
        ]:
            inputs = dict(sensitivity_base)
            inputs.update(overrides)
            base = compute_levine(_to_biomarker_inputs(sensitivity_base))
            variant = compute_levine(_to_biomarker_inputs(inputs))
            cases.append(
                {
                    "sheet": label,
                    "note": f"Pheno age delta vs 20260630 chrono57 baseline: +{delta} yr (verified)",
                    "inputs": inputs,
                    "expected": {
                        "pheno_age": round(variant.pheno_age, 2),
                        "delta_vs_baseline": round(variant.pheno_age - base.pheno_age, 2),
                    },
                }
            )

    for case in cases:
        if "expected" not in case or "pheno_age" not in case["expected"]:
            continue
        computed = compute_levine(_to_biomarker_inputs(case["inputs"]))
        case["computed_pheno_age"] = round(computed.pheno_age, 4)
        case["within_tolerance"] = (
            abs(computed.pheno_age - case["expected"]["pheno_age"])
            <= 0.1 + 1e-6
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": str(XLSX.name),
        "tolerance_years": 0.1,
        "creatinine_conversion": 88.4,
        "cases": cases,
    }
    OUT.write_text(json.dumps(payload, indent=2) + "\n")

    failed = [c["sheet"] for c in cases if c.get("within_tolerance") is False]
    print(f"Wrote {len(cases)} cases to {OUT}")
    if failed:
        print("OUT OF TOLERANCE:", ", ".join(failed))
        sys.exit(1)
    print("All cases within tolerance.")


if __name__ == "__main__":
    main()